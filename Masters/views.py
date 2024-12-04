import json
import pydoc
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render,redirect
from django.contrib.auth import authenticate, login ,logout,get_user_model
from Account.forms import RegistrationForm
from Account.models import *
import Db 
import bcrypt
from django.contrib.auth.decorators import login_required
from Masters.serializers import ScRosterSerializer
from Notification.models import notification_log
from Masters.models import site_master as sit, company_master as com
from Notification.serializers import NotificationSerializer
from PSN.encryption import *
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph
from Account.utils import decrypt_email, encrypt_email
import requests
import traceback
import pandas as pd
from django.core.files.storage import FileSystemStorage
from django.conf import settings
from django.contrib import messages
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
import calendar
from datetime import datetime, timedelta
from django.utils import timezone
from datetime import timedelta
from django.db.models import Q, Count

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework_simplejwt.tokens import AccessToken
from django.utils import timezone
from .models import Log, sc_roster, sc_employee_master, CustomUser 

@login_required
def masters(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor=m.cursor()
    pre_url = request.META.get('HTTP_REFERER')
    header, data = [], []
    entity, type, name = '', '', ''
    global user
    user  = request.session.get('user_id', '')
    try:
         
        if request.method=="GET":
            entity = request.GET.get('entity', '')
            type = request.GET.get('type', '')
            cursor.callproc("stp_get_masters",[entity,type,'name',user])
            for result in cursor.stored_results():
                datalist1 = list(result.fetchall())
            name = datalist1[0][0]
            cursor.callproc("stp_get_masters", [entity, type, 'header',user])
            for result in cursor.stored_results():
                header = list(result.fetchall())
            cursor.callproc("stp_get_masters",[entity,type,'data',user])
            for result in cursor.stored_results():
                if (entity == 'em' or entity == 'sm' or entity == 'cm' or entity == 'menu' or entity == 'user') and type !='err': 
                    data = []
                    rows = result.fetchall()
                    for row in rows:
                        encrypted_id = encrypt_parameter(str(row[0]))
                        data.append((encrypted_id,) + row[1:])
                else: data = list(result.fetchall())

            cursor.callproc("stp_companyfilter",[user])
            for result in cursor.stored_results():
                company_names = list(result.fetchall())
            cursor.callproc("stp_worsitefilter",[user])
            for result in cursor.stored_results():
                site_name = list(result.fetchall())
            if entity == 'r' and type == 'i':
                cursor.callproc("stp_get_assigned_company",[user])
                for result in cursor.stored_results():
                    company_names = list(result.fetchall())
                cursor.callproc("stp_worsitefilter",[user])
                for result in cursor.stored_results():
                    site_name = list(result.fetchall())
            if entity == 'r' and type == 'ed':
                month_year =str(request.GET.get('month', ''))
                if month_year == '':
                    year,month = '',''
                else: year,month = month_year.split('-')
                employee_id = request.GET.get('empid', '')
                cursor.callproc("stp_get_edit_roster",[employee_id,month,year,'1'])
                for result in cursor.stored_results():
                    data = list(result.fetchall())
                cursor.callproc("stp_get_edit_roster",[employee_id,month,year,'2'])
                for result in cursor.stored_results():
                    header = list(result.fetchall())
            if entity == 'urm' and (type == 'acu' or type == 'acr'):
                cursor.callproc("stp_get_access_control",[entity,type])
                for result in cursor.stored_results():
                    header = list(result.fetchall())
                cursor.callproc("stp_get_access_control",[entity,'comp'])
                for result in cursor.stored_results():
                    company_names = list(result.fetchall())
                cursor.callproc("stp_get_access_control",[entity,'site'])
                for result in cursor.stored_results():
                    data = list(result.fetchall())
                
        if request.method=="POST":
            entity = request.POST.get('entity', '')
            type = request.POST.get('type', '')
            if entity == 'r' and type == 'ed':
                ids = request.POST.getlist('ids[]', '')
                shifts = request.POST.getlist('shifts[]', '')
                for id,shift in zip(ids, shifts):
                    cursor.callproc("stp_post_roster",[id,shift])
                    for result in cursor.stored_results():
                        datalist = list(result.fetchall())
                if datalist[0][0] == "success":
                    messages.success(request, 'Data updated successfully !')
            elif entity == 'urm' and (type == 'acu' or type == 'acr'):
                
            #     created_by = request.session.get('user_id', '')
            #     ur = request.POST.get('ur', '')
            #     selected_company_ids = list(map(int, request.POST.getlist('company_id')))
            #     selected_worksites  = request.POST.getlist('worksite')
            #     company_worksite_map = {}
                
            #     if not selected_company_ids or not selected_worksites:
            #         messages.error(request, 'Company or worksite data is missing!')
            #         return redirect(f'/masters?entity={entity}&type=urm')
            #     if type not in ['acu', 'acr'] or not ur:
            #         messages.error(request, 'Invalid data received.')
            #         return redirect(f'/masters?entity={entity}&type=urm')
                
            #     cursor.callproc("stp_get_company_worksite",[",".join(request.POST.getlist('company_id'))])
            #     for result in cursor.stored_results():
            #         company_worksites  = list(result.fetchall())
                    
            #     for company_id, worksite_name in company_worksites:
            #         if company_id not in company_worksite_map:
            #             company_worksite_map[company_id] = []
            #         company_worksite_map[company_id].append(worksite_name)
                
            #     filtered_combinations = []
            #     for company_id in selected_company_ids:
            #         valid_worksites = company_worksite_map.get(company_id, [])
            #         # Filter worksites that were actually selected by the user
            #         selected_valid_worksites = [ws for ws in selected_worksites if ws in valid_worksites]
            #         filtered_combinations.extend([(company_id, ws) for ws in selected_valid_worksites])
                    
            #     cursor.callproc("stp_delete_access_control",[type,ur])
            #     r=''
            #     for company_id, worksite in filtered_combinations:
            #         cursor.callproc("stp_post_access_control",[type,ur,company_id,worksite,created_by])
            #         for result in cursor.stored_results():
            #                 r = list(result.fetchall())
            #     type='urm'
            #     if r[0][0] == "success":
            #         messages.success(request, 'Data updated successfully !')
                
            # else : messages.error(request, 'Oops...! Something went wrong!')
                # created_by = request.session.get('user_id', '')
                # ur = request.POST.get('ur', '')
                # selected_worksite = request.POST.getlist('worksite', [])
                # company_worksite_map = {}

                # # Validate input
                # if not selected_worksite:
                #     messages.error(request, 'Worksite data is missing!')
                #     return redirect(f'/masters?entity={entity}&type=urm')

                # if type not in ['acu', 'acr'] or not ur:
                #     messages.error(request, 'Invalid data received.')
                #     return redirect(f'/masters?entity={entity}&type=urm')

                # selected_worksite_pairs = [
                #     tuple(ws.split(" - ", 1)) for ws in selected_worksite if " - " in ws
                # ]
                # if not selected_worksite_pairs:
                #     messages.error(request, 'Invalid worksite format. Expected "Company Name - Worksite Name".')
                #     return redirect(f'/masters?entity={entity}&type=urm')
                # valid_combinations = []
                # for company_name, worksite_name in selected_worksite_pairs:
                #     try:
                #         company = com.objects.get(company_name=company_name)
                #         company_id = company.company_id
                #     except com.DoesNotExist:
                #         messages.error(request, f'Company "{company_name}" does not exist.')
                #         continue
                #     # Check if the worksite exists for the company in SiteMaster
                #     if sit.objects.filter(company_id=company_id, site_name=worksite_name).exists():
                #         valid_combinations.append((company_id, worksite_name))
                #     else:
                #         messages.error(request, f'Worksite "{worksite_name}" does not exist for company "{company_name}".')  # Assuming first column is company_id
                #         # Check if worksite exists for the company in site_master
                #         cursor.callproc("stp_get_company_worksite", [company_id])
                #         validation_result = [row for result in cursor.stored_results() for row in result.fetchall()]
                #         if validation_result and validation_result[0][0] == 'valid':
                #             valid_combinations.append((company_id, worksite_name))
                #     if not valid_combinations:
                #         messages.error(request, 'No valid company-worksite combinations found.')
                #         return redirect(f'/masters?entity={entity}&type=urm')
                #     # Remove existing mappings
                #     cursor.callproc("stp_delete_access_control", [type, ur])
                #     # Insert valid combinations into user_role_map
                #     for company_id, worksite_name in valid_combinations:
                #         cursor.callproc("stp_post_access_control", [type, ur, company_id, worksite_name, created_by])
                #     messages.success(request, 'Data updated successfully!')

                try:
                    created_by = request.session.get('user_id', '')
                    ur = request.POST.get('ur', '')
                    selected_worksite = request.POST.getlist('worksite', [])
                    company_worksite_map = {}

                    # Validate input
                    if not selected_worksite:
                        messages.error(request, 'Worksite data is missing!')
                        return redirect(f'/masters?entity={entity}&type=urm')

                    if type not in ['acu', 'acr'] or not ur:
                        messages.error(request, 'Invalid data received.')
                        return redirect(f'/masters?entity={entity}&type=urm')

                    # Parse selected worksites into company-worksite pairs
                    selected_worksite_pairs = [
                        tuple(ws.split(" - ", 1)) for ws in selected_worksite if " - " in ws
                    ]
                    if not selected_worksite_pairs:
                        messages.error(request, 'Invalid worksite format. Expected "Company Name - Worksite Name".')
                        return redirect(f'/masters?entity={entity}&type=urm')

                    valid_combinations = []
                    for company_name, worksite_name in selected_worksite_pairs:
                        try:
                            # Fetch company_id using ORM
                            company = com.objects.get(company_name=company_name)
                            company_id = company.company_id
                        except com.DoesNotExist:
                            messages.error(request, f'Company "{company_name}" does not exist.')
                            continue

                        # Check if the worksite exists for the company in SiteMaster
                        if sit.objects.filter(company_id=company_id, site_name=worksite_name).exists():
                            valid_combinations.append((company_id, worksite_name))
                        else:
                            messages.error(request, f'Worksite "{worksite_name}" does not exist for company "{company_name}".')

                    if not valid_combinations:
                        messages.error(request, 'No valid company-worksite combinations found.')

                    # Remove existing mappings
                    cursor.callproc("stp_delete_access_control", [type, ur])

                    # Insert new mappings
                    insertion_status = "failure"
                    for company_id, worksite_name in valid_combinations:
                        cursor.callproc("stp_post_access_control", [type, ur, company_id, worksite_name, created_by])
                        for result in cursor.stored_results():
                            r = list(result.fetchall())
                            if r and r[0][0] == "success":
                                insertion_status = "success"

                    # Redirect based on insertion status
                    if insertion_status == "success":
                        messages.success(request, 'Data updated successfully!')
                    else:
                        messages.error(request, 'Oops...! Something went wrong!')
                except Exception as e:
                    tb = traceback.extract_tb(e.__traceback__)
                    fun = tb[0].name
                    cursor.callproc("stp_error_log",[fun,str(e),user])  
                    messages.error(request, 'Oops...! Something went wrong!')

                    
            else : messages.error(request, 'Oops...! Something went wrong!')

    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        cursor.callproc("stp_error_log",[fun,str(e),user])  
        messages.error(request, 'Oops...! Something went wrong!')
    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()
        if request.method=="GET":
            return render(request,'Master/index.html', {'entity':entity,'type':type,'name':name,'header':header,'company_names':company_names,'data':data,'pre_url':pre_url,'site_name':site_name})
        elif request.method=="POST":  
            if entity == 'r':
                new_url = f'/masters?entity={entity}&type=i'
                return redirect(new_url)
            elif entity == 'urm':
                new_url = f'/masters?entity={entity}&type=urm'
                return redirect(new_url)
            else:
                new_url = f'/masters?entity={entity}&type={type}'
                return redirect(new_url)
         
        
def gen_roster_xlsx_col(columns,month_input):
    year, month = map(int, month_input.split('-'))
    _, num_days = calendar.monthrange(year, month)
    date_columns = [(datetime(year, month, day)).strftime('%d-%m-%Y') for day in range(1, num_days + 1)]
    columns.extend(date_columns)
    return columns
        
def sample_xlsx(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor=m.cursor()
    pre_url = request.META.get('HTTP_REFERER')
    response =''
    global user
    user  = request.session.get('user_id', '')
    try:
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Sample Format'
        columns = []
        if request.method=="GET":
            entity = request.GET.get('entity', '')
            type = request.GET.get('type', '')
        if request.method=="POST":
            entity = request.POST.get('entity', '')
            type = request.POST.get('type', '')
        file_name = {'em': 'Employee Master','sm': 'Worksite Master','cm': 'Company Master','r': 'Roster'}[entity]
        cursor.callproc("stp_get_masters", [entity, type, 'sample_xlsx',user])
        for result in cursor.stored_results():
            columns = [col[0] for col in result.fetchall()]
        # columns = ['Column 1', 'Column 2', 'Column 3']
        if entity == "r":
            month = request.POST.get('month', '')
            columns = gen_roster_xlsx_col(columns,month)

        black_border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )
        
        for col_num, header in enumerate(columns, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.border = black_border
        
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  
            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
                    
            adjusted_width = max_length + 2 
            sheet.column_dimensions[column].width = adjusted_width  
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="' + str(file_name) +" "+str(datetime.now().strftime("%d-%m-%Y")) + '.xlsx"'
        workbook.save(response)
    
    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        cursor.callproc("stp_error_log",[fun,str(e),user])  
        messages.error(request, 'Oops...! Something went wrong!')
    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()
        return response      

@login_required  
def roster_upload(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()
    global user
    user  = request.session.get('user_id', '')
    if request.method == 'POST' and request.FILES.get('roster_file'):
        try:
            excel_file = request.FILES['roster_file']
            file_name = excel_file.name
            df = pd.read_excel(excel_file)

            entity = request.POST.get('entity', '')
            type = request.POST.get('type', '')
            company_id = request.POST.get('company_id', '')
            month_input  =str(request.POST.get('month_year', ''))
            total_rows = len(df)
            update_count = error_count = success_count = 0
            checksum_id = None
            worksites = []

            if entity == 'r':
                year, month = map(int, month_input.split('-'))
                _, num_days = calendar.monthrange(year, month)
                date_columns = [(datetime(year, month, day)).strftime('%d-%m-%Y') for day in range(1, num_days + 1)]
                cursor.callproc("stp_get_masters", [entity, type, 'sample_xlsx',user])
                for result in cursor.stored_results():
                    start_columns = [col[0] for col in result.fetchall()]

                if not all(col in df.columns for col in start_columns + date_columns):
                    messages.error(request, 'Oops...! The uploaded Excel file does not contain the required columns.!')
                    return redirect(f'/masters?entity={entity}&type={type}')
                
                cursor.callproc('stp_insert_checksum', ('roster',company_id,month,year,file_name))
                for result in cursor.stored_results():
                    c = list(result.fetchall())
                checksum_id = c[0][0]
                
                for index,row in df.iterrows():
                    employee_id = row.get('Employee Id', '')
                    employee_name = row.get('Employee Name', '')
                    worksite  = row.get('Worksite', '')
                    
                    for date_col in date_columns:
                        shift_date = datetime.strptime(date_col, '%d-%m-%Y').date()
                        shift_time = row.get(date_col) 
                        if pd.isna(shift_time):
                            shift_time = None
                        params = (str(employee_id),employee_name,int(company_id),worksite,shift_date,shift_time,checksum_id,user)
                        cursor.callproc('stp_insert_roster', params)
                        for result in cursor.stored_results():
                            r = list(result.fetchall())
                        if r[0][0] not in ("success", "updated"):
                            if worksite not in worksites:
                                worksites.append(worksite)
                            error_message = str(r[0][0])
                            error_params = ('roster', company_id,worksite,file_name,shift_date,error_message,checksum_id)
                            cursor.callproc('stp_insert_error_log', error_params)
                            messages.error(request, "Errors occurred during upload. Please check error logs.")
                    if r[0][0] == "success": success_count += 1
                    elif r[0][0] == "updated": update_count += 1  
                    else: error_count += 1
                checksum_msg = f"Total Rows Processed: {total_rows}, Successful Entries: {success_count}, Updates: {update_count}, Errors: {error_count}"
                cursor.callproc('stp_update_checksum', ('roster',company_id,', '.join(worksites),month,year,file_name,checksum_msg,error_count,update_count,checksum_id))
                if error_count == 0 and update_count == 0 and success_count > 0:
                    messages.success(request, f"All data uploaded successfully!.")
                elif error_count == 0 and success_count == 0 and update_count > 0:
                    messages.warning(request, f"All data updated successfully!.")
                else:
                    messages.warning(request, f"The upload processed {total_rows} rows, resulting in {success_count} successful entries, {update_count} updates, and {error_count} errors; please check the error logs for details.")
                    
        except Exception as e:
            tb = traceback.extract_tb(e.__traceback__)
            fun = tb[0].name
            cursor.callproc("stp_error_log", [fun, str(e), user])  
            messages.error(request, 'Oops...! Something went wrong!')
            m.commit()   

        finally:
            cursor.close()
            m.close()
            Db.closeConnection()
            return redirect(f'/masters?entity={entity}&type={type}')     
        
@login_required        
def site_master(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor=m.cursor()
    global user
    user  = request.session.get('user_id', '')
    try:
        
        if request.method == "GET":
            cursor.callproc("stp_get_roster_type")
            for result in cursor.stored_results():
                roster_types = list(result.fetchall())
                # Call stored procedure to get company names
            cursor.callproc("stp_get_graph_dropdown", [user,'company'])
            for result in cursor.stored_results():
                company_names = list(result.fetchall())

        cursor.callproc("stp_get_graph_dropdown", [user,'site'])
        for result in cursor.stored_results():
            site_names = list(result.fetchall())
            site_id = request.GET.get('site_id', '')
            if site_id == "0":
                if request.method == "GET":
                    context = {'company_names': company_names, 'roster_type': roster_types,'site_id':site_id}

            else:
                site_id1 = request.GET.get('site_id', '')
                site_id = decrypt_parameter(site_id1)
                cursor.callproc("stp_edit_site_master", (site_id,)) 
                for result in cursor.stored_results():
                    data = result.fetchall()[0]  
                    context = {
                        'roster_types':roster_types,
                        'company_names':company_names,
                        'site_id' : data[0],
                        'site_name': data[1],
                        'site_address': data[2],
                        'pincode': data[3],
                        'contact_person_name': data[4],
                        'contact_person_email': data[5], 
                        'contact_person_mobile_no': data[6],
                        'is_active':data[7],
                        'no_of_days': data[8],               
                        'notification_time': data[9],
                        'reminder_time': data[10],
                        'company_name' :data[11],
                        'roster_type': data[13]
                    }

        if request.method == "POST":
            siteId = request.POST.get('site_id', '')
            if siteId == "0":
                response_data = {"status": "fail"}
                
                siteName = request.POST.get('siteName', '')
                siteAddress = request.POST.get('siteAddress', '')
                pincode = request.POST.get('pincode', '')
                contactPersonName = request.POST.get('contactPersonName', '')
                contactPersonEmail = request.POST.get('contactPersonEmail', '')
                contactPersonMobileNo = request.POST.get('Number', '')  
                # is_active = request.POST.get('status_value', '') 
                # noOfDays = request.POST.get('FieldDays', '')  
                # notificationTime = request.POST.get('notificationTime', '')
                # ReminderTime = request.POST.get('ReminderTime', '')
                companyId = request.POST.get('company_id', '')  
                # rosterType = request.POST.get('roster_type', '')
               
                params = [
                    siteName, 
                    siteAddress, 
                    pincode, 
                    contactPersonName, 
                    contactPersonEmail, 
                    contactPersonMobileNo, 
                    # is_active,
                    # noOfDays, 
                    # notificationTime, 
                    # ReminderTime, 
                    companyId
                    # rosterType
                ]
                
                cursor.callproc("stp_insert_site_master", params)
                for result in cursor.stored_results():
                        datalist = list(result.fetchall())
                if datalist[0][0] == "success":
                    messages.success(request, 'Data successfully entered !')
                else: messages.error(request, datalist[0][0])
            else:
                if request.method == "POST" :
                    siteId = request.POST.get('site_id', '')
                    siteName = request.POST.get('siteName', '')
                    siteAddress = request.POST.get('siteAddress', '')
                    pincode = request.POST.get('pincode', '')
                    contactPersonName = request.POST.get('contactPersonName', '')
                    contactPersonEmail = request.POST.get('contactPersonEmail', '')
                    contactPersonMobileNo = request.POST.get('Number', '')  
                    # noOfDays = request.POST.get('FieldDays', '') 
                    isActive = request.POST.get('status_value', '')
                    # notificationTime = request.POST.get('notificationTime', '')
                    # ReminderTime = request.POST.get('ReminderTime', '')
                    CompanyId = request.POST.get('company_id', '')
                    # Rostertype = request.POST.get('roster_type', '')
                    
                    params = [siteId,siteName,siteAddress,pincode,contactPersonName,contactPersonEmail,
                                        contactPersonMobileNo,isActive,CompanyId]
                    cursor.callproc("stp_update_site_master",params) 
                    messages.success(request, "Data updated successfully...!")

    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        cursor.callproc("stp_error_log", [fun, str(e), user])  
        messages.error(request, 'Oops...! Something went wrong!')
    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()
            
        if request.method=="GET":
            return render(request, "Master/site_master.html", context)
        elif request.method=="POST":  
            return redirect( f'/masters?entity=sm&type=i')
        
@login_required      
def company_master(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor=m.cursor()
    global user
    user  = request.session.get('user_id', '')
    try:
        
        if request.method == "GET":
        
            company_id = request.GET.get('company_id', '')
            if company_id == "0":
                if request.method == "GET":
                    context = {'company_id':company_id}
            else:
                company_id1 = request.GET.get('company_id', '')
                company_id= decrypt_parameter(company_id1)
                cursor.callproc("stp_edit_company_master", (company_id,))  # Note the comma to make it a tuple
                for result in cursor.stored_results():
                    data = result.fetchall()[0]  
                        
                    context = {
                        'company_id' : data[0],
                        'company_name': data[1],
                        'company_address': data[2],
                        'pincode': data[3],
                        'contact_person_name': data[4],
                        'contact_person_email': data[5], 
                        'contact_person_mobile_no': data[6],
                        'is_active':data[7]
                    }

        if request.method == "POST" :
            company_id = request.POST.get('company_id', '')
            if company_id == '0':
                response_data = {"status": "fail"}
                company_name = request.POST.get('company_name', '')
                company_address = request.POST.get('company_address', '')
                pincode = request.POST.get('pincode', '')
                contact_person_name = request.POST.get('contact_person_name', '')
                contact_person_email = request.POST.get('contact_person_email', '')
                contact_person_mobile_no = request.POST.get('contact_person_mobile_no', '') 
                # is_active = request.POST.get('status_value', '') 
                params = [
                    company_name, 
                    company_address, 
                    pincode, 
                    contact_person_name,
                    contact_person_email,
                    contact_person_mobile_no
                    # is_active
                ]
                cursor.callproc("stp_insert_company_master", params)
                for result in cursor.stored_results():
                        datalist = list(result.fetchall())
                if datalist[0][0] == "success":
                    messages.success(request, 'Data successfully entered !')
                else: messages.error(request, datalist[0][0])
            else :
                company_id = request.POST.get('company_id', '')
                company_name = request.POST.get('company_name', '')
                company_address = request.POST.get('company_address', '')
                pincode = request.POST.get('pincode', '')
                contact_person_name = request.POST.get('contact_person_name', '')
                contact_person_email = request.POST.get('contact_person_email', '')
                contact_person_mobile_no = request.POST.get('contact_person_mobile_no', '') 
                is_active = request.POST.get('status_value', '') 
                   
                params = [company_id,company_name,company_address,pincode,contact_person_name,contact_person_email,
                                            contact_person_mobile_no,is_active]    
                cursor.callproc("stp_update_company_master",params) 
                messages.success(request, "Data updated successfully ...!")
                
    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        cursor.callproc("stp_error_log", [fun, str(e), user])  
        messages.error(request, 'Oops...! Something went wrong!')
    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()

        encrypted_id = encrypt_parameter(company_id)
            
        if request.method=="GET":
            return render(request, "Master/company_master.html", context)
        elif request.method == "POST":
            return redirect(f'/masters?entity=cm&type=i')

# @login_required        
# def employee_master(request):
#     Db.closeConnection()
#     m = Db.get_connection()
#     cursor=m.cursor()
#     global user
#     user  = request.session.get('user_id', '')
#     try:
        
#         if request.method == "GET":
#             id = request.GET.get('id', '')
            
#             cursor.callproc("stp_get_company_names")
#             for result in cursor.stored_results():
#                 company_names = list(result.fetchall())
#             cursor.callproc("stp_get_employee_status")
#             for result in cursor.stored_results():
#                 employee_status = list(result.fetchall())
#             cursor.callproc("stp_get_dropdown_values",('site',))
#             for result in cursor.stored_results():
#                 site_name = list(result.fetchall())
#             if id == "0":
#                 if request.method == "GET":
#                     context = {'id':id, 'employee_status':employee_status, 'employee_status_id': '','site_name':site_name}

#             else:
#                 id1 = request.GET.get('id', '')
#                 id = decrypt_parameter(id1)
#                 cursor.callproc("stp_edit_employee_master", (id,))
#                 for result in cursor.stored_results():
#                     data = result.fetchall()[0]  
#                     context = {
#                         'site_name':site_name,
#                         'employee_status':employee_status,
#                         'id':data[0],
#                         'employee_id' : data[1],
#                         'employee_name': data[2],
#                         'mobile_no': data[3],
#                         'site_name_value': data[4],
#                         'employee_status_id': data[5],
#                         'is_active': data[6]
#                     }

#         if request.method == "POST" :
#             id = request.POST.get('id', '')
#             if id == '0':

#                 employeeId = request.POST.get('employee_id', '')
#                 employeeName = request.POST.get('employee_name', '')
#                 mobileNo = request.POST.get('mobile_no', '')
#                 site_name = request.POST.get('site_name', '')
#                 # employeeStatus = request.POST.get('employee_status_name', '')
#                 # activebtn = request.POST.get('status_value', '')

#                 params = [
#                     employeeId, 
#                     employeeName, 
#                     mobileNo, 
#                     site_name
#                     # employeeStatus,
#                     # activebtn
#                 ]
                
#                 cursor.callproc("stp_insert_employee_master", params)
#                 for result in cursor.stored_results():
#                         datalist = list(result.fetchall())
#                 if datalist[0][0] == "success":
#                     messages.success(request, 'Data successfully entered !')
#                 else: messages.error(request, datalist[0][0])
#             else:
#                 id = request.POST.get('id', '')
#                 employee_id = request.POST.get('employee_id', '')
#                 employee_name = request.POST.get('employee_name', '')
#                 mobile_no = request.POST.get('mobile_no', '')
#                 site_name = request.POST.get('site_name', '')
#                 employee_status = request.POST.get('employee_status_name', '')
#                 is_active = request.POST.get('status_value', '')  
                            
#                 params = [id,employee_id,employee_name,mobile_no,site_name,employee_status,is_active]    
#                 cursor.callproc("stp_update_employee_master",params) 
#                 messages.success(request, "Data successfully Updated!")

#     except Exception as e:
#         tb = traceback.extract_tb(e.__traceback__)
#         fun = tb[0].name
#         cursor.callproc("stp_error_log", [fun, str(e), user])  
#         messages.error(request, 'Oops...! Something went wrong!')
#     finally:
#         cursor.close()
#         m.commit()
#         m.close()
#         Db.closeConnection()
#         if request.method=="GET":
#             return render(request, "Master/employee_master.html", context)
#         elif request.method=="POST":  
#             return redirect(f'/masters?entity=em&type=i')
@login_required        
def employee_master(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor=m.cursor()
    global user
    user  = request.session.get('user_id', '')
    try:
        
        if request.method == "GET":
            id = request.GET.get('id', '')
            
            cursor.callproc("stp_get_graph_dropdown", [user,'company'])
            for result in cursor.stored_results():
                company_names = list(result.fetchall())

            cursor.callproc("stp_get_employee_status")
            for result in cursor.stored_results():
                employee_status = list(result.fetchall())
            if id == "0":
                cursor.callproc("stp_get_graph_dropdown", [user,'site'])
                for result in cursor.stored_results():
                    site_name = list(result.fetchall())
            else:
                id1 = decrypt_parameter(id)
                cursor.callproc("stp_get_employee_master_worksite",(id1,))
                for result in cursor.stored_results():
                    site_name = list(result.fetchall())
            if id == "0":
                if request.method == "GET":
                    context = {'id':id, 'employee_status':employee_status,'company_names': company_names, 'employee_status_id': '','site_name':site_name}

            else:
                id1 = request.GET.get('id', '')
                id = decrypt_parameter(id1)
                cursor.callproc("stp_edit_employee_master", (id,))
                for result in cursor.stored_results():
                    data = result.fetchall()[0]  
                    context = {
                        'site_name': site_name,
                        'employee_status': employee_status,
                        'company_names': company_names,
                        'id': data[0],
                        'employee_id': data[1],
                        'employee_name': data[2],
                        'mobile_no': data[3],
                        'site_name_value': data[4],
                        'is_active': data[7],
                        'company_name': data[5],  # Ensure this holds the company ID, not the name
                        'employee_status_id': data[6]
                    }


        if request.method == "POST" :
            id = request.POST.get('id', '')
            if id == '0':

                employeeId = request.POST.get('employee_id', '')
                employeeName = request.POST.get('employee_name', '')
                mobileNo = request.POST.get('mobile_no', '')
                site_name = request.POST.get('site_name', '')
                CompanyId = request.POST.get('company_id', '')
                # employeeStatus = request.POST.get('employee_status_name', '')
                # activebtn = request.POST.get('status_value', '')

                params = [
                    employeeId, 
                    employeeName, 
                    mobileNo, 
                    site_name,
                    CompanyId
                    # employeeStatus,
                    # activebtn
                ]
                
                cursor.callproc("stp_insert_employee_master", params)
                for result in cursor.stored_results():
                        datalist = list(result.fetchall())
                if datalist[0][0] == "success":
                    messages.success(request, 'Data successfully entered !')
                else: messages.error(request, datalist[0][0])
            else:
                id = request.POST.get('id', '')
                employee_id = request.POST.get('employee_id', '')
                employee_name = request.POST.get('employee_name', '')
                mobile_no = request.POST.get('mobile_no', '')
                site_name = request.POST.get('site_name', '')
                CompanyId = request.POST.get('company_id', '')
                employee_status = request.POST.get('employee_status_name', '')
                is_active = request.POST.get('status_value', '')  
                            
                params = [id,employee_id,employee_name,mobile_no,site_name,CompanyId,employee_status,is_active]    
                cursor.callproc("stp_update_employee_master",params) 
                messages.success(request, "Data successfully Updated!")

    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        cursor.callproc("stp_error_log", [fun, str(e), user])  
        messages.error(request, 'Oops...! Something went wrong!')
    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()
        if request.method=="GET":
            return render(request, "Master/employee_master.html", context)
        elif request.method=="POST":  
            return redirect(f'/masters?entity=em&type=i')

@login_required  
def upload_excel(request):

    if request.method == 'POST' and request.FILES.get('excelFile'):
        excel_file = request.FILES['excelFile']
        file_name = excel_file.name
        df = pd.read_excel(excel_file)
        total_rows = len(df)
        update_count = error_count = success_count = 0
        checksum_id = None
        r=None
        global user
        user  = request.session.get('user_id', '')
        try:
            Db.closeConnection()
            m = Db.get_connection()
            cursor = m.cursor()
            entity = request.POST.get('entity', '')
            type = request.POST.get('type', '')
            company_id1 = request.POST.get('company_id', None)
            cursor.callproc("stp_get_masters", [entity, type, 'sample_xlsx',user])
            for result in cursor.stored_results():
                columns = [col[0] for col in result.fetchall()]
            if not all(col in df.columns for col in columns):
                messages.error(request, 'Oops...! The uploaded Excel file does not contain the required columns.!')
                return redirect(f'/masters?entity={entity}&type={type}')
            upload_for = {'em': 'employee master','sm': 'site master','cm': 'company master','r': 'roster'}[entity]
            cursor.callproc('stp_insert_checksum', (upload_for,company_id1,str(datetime.now().month),str(datetime.now().year),file_name))
            for result in cursor.stored_results():
                c = list(result.fetchall())
            checksum_id = c[0][0]

            # if entity == 'em':
            #     for index,row in df.iterrows():
            #         params = tuple(str(row.get(column, '')) for column in columns)
            #         cursor.callproc('stp_insert_employee_master', params)
            #         for result in cursor.stored_results():
            #                 r = list(result.fetchall())
            #         if r[0][0] not in ("success", "updated"):
            #             cursor.callproc('stp_insert_error_log', [upload_for, company_id,'',file_name,datetime.now().date(),str(r[0][0]),checksum_id])
            #         if r[0][0] == "success": success_count += 1 
            #         elif r[0][0] == "updated": update_count += 1  
            #         else: error_count += 1
            if entity == 'em':
                for index,row in df.iterrows():
                    params = tuple(str(row.get(column, '')) for column in columns)
    
                    # Get the company instance and add company_id to params
                    params += (str(company_id1),) # Assuming company_id is an integer
                    
                    # Debugging: Print params to verify
                    print("Params being passed to stored procedure:", params)
                    
                    # Call the stored procedure
                    cursor.callproc('stp_insert_employee_master', params)
                    for result in cursor.stored_results():
                            r = list(result.fetchall())
                    if r[0][0] not in ("success", "updated"):
                        cursor.callproc('stp_insert_error_log', [upload_for, company_id1,'',file_name,datetime.now().date(),str(r[0][0]),checksum_id])
                    if r[0][0] == "success": success_count += 1 
                    elif r[0][0] == "updated": update_count += 1  
                    else: error_count += 1
            elif entity == 'sm':
                for index,row in df.iterrows():
                    params = tuple(str(row.get(column, '')) for column in columns)
                    params += (str(company_id1),)
                    cursor.callproc('stp_insert_site_master', params)
                    for result in cursor.stored_results():
                            r = list(result.fetchall())
                    if r[0][0] not in ("success", "updated"):
                        cursor.callproc('stp_insert_error_log', [upload_for, company_id1,'',file_name,datetime.now().date(),str(r[0][0]),checksum_id])
                    if r[0][0] == "success": success_count += 1 
                    elif r[0][0] == "updated": update_count += 1  
                    else: error_count += 1
            elif entity == 'cm':
                for index,row in df.iterrows():
                    params = tuple(str(row.get(column, '')) for column in columns)
                    cursor.callproc('stp_insert_company_master', params)
                    for result in cursor.stored_results():
                            r = list(result.fetchall())
                    if r[0][0] not in ("success", "updated"):
                        cursor.callproc('stp_insert_error_log', [upload_for, company_id1,'',file_name,datetime.now().date(),str(r[0][0]),checksum_id])
                    if r[0][0] == "success": success_count += 1 
                    elif r[0][0] == "updated": update_count += 1  
                    else: error_count += 1
            checksum_msg = f"Total Rows Processed: {total_rows}, Successful Entries: {success_count}" f"{f', Updates: {update_count}' if update_count > 0 else ''}" f"{f', Errors: {error_count}' if error_count > 0 else ''}"
            cursor.callproc('stp_update_checksum', (upload_for,company_id1,'',str(datetime.now().month),str(datetime.now().year),file_name,checksum_msg,error_count,update_count,checksum_id))
            if error_count == 0 and update_count == 0 and success_count > 0:
                messages.success(request, f"All data uploaded successfully!.")
            elif error_count == 0 and success_count == 0 and update_count > 0:
                messages.warning(request, f"All data updated successfully!.")
            else:messages.warning(request, f"The upload processed {total_rows} rows, resulting in {success_count} successful entries"  f"{f', {update_count} updates' if update_count > 0 else ''}" f", and {error_count} errors; please check the error logs for details.")
                   
        except Exception as e:
            tb = traceback.extract_tb(e.__traceback__)
            fun = tb[0].name
            cursor.callproc("stp_error_log", [fun, str(e), user])  
            messages.error(request, 'Oops...! Something went wrong!')
            m.commit()   
        finally:
            cursor.close()
            m.close()
            Db.closeConnection()
            return redirect(f'/masters?entity={entity}&type=i')

# def get_access_control(request):
#     Db.closeConnection()
#     m = Db.get_connection()
#     cursor=m.cursor()
#     company = []
#     worksite = []
#     global user
#     user  = request.session.get('user_id', '')
#     try:
#         if request.method == "POST":
#             type = request.POST.get('type','')
#             ur = request.POST.get('ur', '')
#             cursor.callproc("stp_get_access_control_val", [type,ur,'company'])
#             for result in cursor.stored_results():
#                 company = list(result.fetchall())
#             cursor.callproc("stp_get_access_control_val", [type,ur,'worksite'])
#             for result in cursor.stored_results():
#                 worksite = list(result.fetchall())
#             if type == 'worksites':
#                 company_id = request.POST.getlist('company_id','')
#                 company_ids = ','.join(company_id)
#                 cursor.callproc("stp_get_access_control_val", [type,company_ids,'worksites'])
#                 for result in cursor.stored_results():
#                     worksite = list(result.fetchall())
                    
#             response = {'result': 'success', 'company': company, 'worksite': worksite}
#         else: response = {'result': 'fail', 'message': 'Invalid request method'}

#     except Exception as e:
#         tb = traceback.extract_tb(e.__traceback__)
#         cursor.callproc("stp_error_log", [tb[0].name, str(e), user])
#         print(f"error: {e}")
#         response = {'result': 'fail', 'message': 'Something went wrong!'}

#     finally:
#         cursor.close()
#         m.close()
#         Db.closeConnection()
#         return JsonResponse(response)

def get_access_control(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor=m.cursor()
    company = []
    worksite = []
    global user
    user  = request.session.get('user_id', '')
    try:
        if request.method == "POST":
            type1 = request.POST.get('type1','')
            type = request.POST.get('type','')
            ur = request.POST.get('ur', '')
            cursor.callproc("stp_get_access_control_val", [type,ur,'company'])
            for result in cursor.stored_results():
                company = list(result.fetchall())
            cursor.callproc("stp_get_access_control_val", [type,ur,'worksite'])
            for result in cursor.stored_results():
                worksite = list(result.fetchall())
            if type1 == 'worksites':
                company_id = request.POST.getlist('company_id','')
                company_ids = ','.join(company_id)
                cursor.callproc("stp_get_access_control_val", [type1,company_ids,'worksites'])
                for result in cursor.stored_results():
                    worksite = list(result.fetchall())
                response = {
                'result': 'success',
                'worksite':worksite,
                'company': company,
                'worksite': worksite,
                }
            else:     
                response = {
                'result': 'success',
                'worksite':worksite,
                'company': company,
                }

        # Return JSON response
            return JsonResponse(response)
        else: response = {'result': 'fail', 'message': 'Invalid request method'}

    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        cursor.callproc("stp_error_log", [tb[0].name, str(e), user])
        print(f"error: {e}")
        response = {'result': 'fail', 'message': 'Something went wrong!'}

    finally:
        cursor.close()
        m.close()
        Db.closeConnection()
        return JsonResponse(response)

class RosterDataAPIView(APIView):
    # Ensure the user is authenticated using JWT
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    def get(self, request):
        # Extract user ID from the JWT token
        user = request.user  # This will get the user from the JWT token

        # Call the function to get the roster data
        roster_data = self.get_roster_data(user.id)
        Log.objects.create(log_text=f"Fetched user by ID: {user.id}")

        return Response(roster_data)

    def get_roster_data(self, user_id):
        # Step 1: Get the user by user_id
        user = CustomUser.objects.get(id=user_id)
        
        # Step 2: Get the phone number of the user
        phone_number = user.phone

        # Step 3: Get the employee_id from sc_employee_master using the phone number
        try:
            employee = sc_employee_master.objects.get(mobile_no=phone_number)
        except sc_employee_master.DoesNotExist:
            return {
                'error': 'Employee not found'
            }
        employee_id = employee.employee_id

        # Step 4: Get the current date and the first date of the current month
        current_date = timezone.now().date()

        # Step 5: Query sc_roster for the current month and categorize the data
        current_roster_qs = sc_roster.objects.filter(
            employee_id=employee_id,
            shift_date__gte=current_date,
            shift_time__isnull=False
        )
        
        current_roster_qsser = ScRosterSerializer(current_roster_qs, many=True)

        previous_roster_qs = sc_roster.objects.filter(
            employee_id=employee_id,
            shift_date__lt=current_date,
            shift_time__isnull=False
            
        )
        previous_roster_qsser = ScRosterSerializer(previous_roster_qs, many=True)

        marked_roster_qs = sc_roster.objects.filter(
            employee_id=employee_id,
            confirmation__isnull=False ,
            shift_time__isnull=False
        )
        marked_roster_qsser = ScRosterSerializer(marked_roster_qs, many=True)

        unmarked_roster_qs = sc_roster.objects.filter(
            employee_id=employee_id,
            confirmation__isnull=True ,
            shift_date__lt=current_date,
            shift_time__isnull=False
        )
        unmarked_roster_qsser = ScRosterSerializer(unmarked_roster_qs, many=True)

        # Count the number of rows in each query set
        current_roster_count = len(current_roster_qsser.data)
        previous_roster_count = len(previous_roster_qsser.data)
        marked_roster_count = len(marked_roster_qsser.data)
        unmarked_roster_count = len(unmarked_roster_qsser.data)

        # Return the counts and the lists
        return {
            'current_roster_count': current_roster_count,
            'current_roster_list': list(current_roster_qsser.data ),  # Using .values() to serialize queryset
            'previous_roster_count': previous_roster_count,
            'previous_roster_list': list(previous_roster_qsser.data),  # Using .values() to serialize queryset
            'marked_roster_count': marked_roster_count,
            'marked_roster_list': list(marked_roster_qsser.data),  # Using .values() to serialize queryset
            'unmarked_roster_count': unmarked_roster_count,
            'unmarked_roster_list': list(unmarked_roster_qsser.data),  # Using .values() to serialize queryset
            'roster_list': list(current_roster_qsser.data)  # Same as Current Roster List
        }

class confirm_schedule(APIView):
    # Ensure the user is authenticated using JWT
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]
    # authentication_classes = []

    def post(self, request):
        try:
            data= request.data
            roster_id = request.data.get('id')
            confirmation = request.data.get('confirmation') == '1'
            user = request.user
            roster = sc_roster.objects.get(id=roster_id)
            roster.confirmation = confirmation
            roster.updated_at = timezone.now()
            roster.confirmation_date = timezone.now()
            roster.updated_by = user
            roster.save()
            ser = ScRosterSerializer(roster)

            return Response({'success': 'Confirmation updated successfully.','data':ser.data,'con':confirmation}, status=200)
        except sc_roster.DoesNotExist:
            return Response({'error': 'Roster not found.'}, status=404)
        except Exception as e:
            return Response({'error': str(e)}, status=500)



class confirm_notification(APIView):
    # Ensure the user is authenticated using JWT
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]
    # authentication_classes = []

    def post(self, request):
        try:
            data= request.data
            notification_id = request.data.get('id')
            confirmation = request.data.get('confirmation') == '1'
            user = request.user
            notification = notification_log.objects.get(id=notification_id)
            notification.notification_opened = timezone.now()
            notification.updated_at = timezone.now()
            notification.updated_by = user
            notification.save()

            return Response({'success': 'Confirmation updated successfully.'}, status=200)
        except notification_log.DoesNotExist:
            return Response({'error': 'notification_log not found.'}, status=404)
        except Exception as e:
            return Response({'error': str(e)}, status=500)
        

# Palavee Attendance Changes

@login_required
def attendance_masters(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor=m.cursor()
    pre_url = request.META.get('HTTP_REFERER')
    header, data = [], []
    entity, type, name = '', '', ''
    global user
    user  = request.session.get('user_id', '')
    user1  = request.session.get('user_id', '')
    user2  = request.session.get('user_id', '')
    try:
         
        if request.method=="GET":
            entity = 'au'
            type = 'err'
            cursor.callproc("stp_get_masters",[entity,type,'name',user])
            for result in cursor.stored_results():
                datalist1 = list(result.fetchall())
            name = datalist1[0][0]
            cursor.callproc("stp_get_masters", [entity, type, 'header',user])
            for result in cursor.stored_results():
                header = list(result.fetchall())
            cursor.callproc("stp_companyfilter",[user1])
            for result in cursor.stored_results():
                company_names = list(result.fetchall())
            cursor.callproc("stp_worsitefilter",[user2])
            for result in cursor.stored_results():
                site_name = list(result.fetchall())
            cursor.callproc("stp_get_attendance_data",[user])
            for result in cursor.stored_results():
                data = list(result.fetchall())
                formatted_data = [
                    {
                      
                        "employee_id": row[0],
                        "employee_name": row[1],
                        "company_id": row[2],
                        "worksite": row[3],
                        "shift_time": row[4],
                        "attendance_in": row[5],
                        "attendance_out": row[6],
                        "id_edit": row[7],
                    }
                    for row in data
                ]
            if entity == 'au' and type == 'i':
                cursor.callproc("stp_get_assigned_company",[user])
                for result in cursor.stored_results():
                    company_names = list(result.fetchall())
                           
    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        cursor.callproc("stp_error_log",[fun,str(e),user])  
        messages.error(request, 'Oops...! Something went wrong!')
    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()
        if request.method=="GET":
            return render(request,'Master/Attendance_upload_index.html', {'entity':entity,'type':type,'name':name,'header':header,'company_names':company_names,'site_name':site_name,'data':formatted_data,'pre_url':pre_url})
        elif request.method=="POST":  
            new_url = f'/masters?entity={entity}&type={type}'
            return redirect(new_url) 
        
 # Import your CompanyMaster model



def attendance_sample(request):
    # Make sure to properly close the DB connection and initialize the cursor if needed
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()

    # Create a new workbook and add a sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Attendance Sample"

    # Create headers for the data in row 1 (Bold)
    sheet["A1"] = "Employee ID"
    sheet["B1"] = "Attendance In (24 hours format)"
    sheet["C1"] = "Attendance Out (24 hours format)"

    # Make headers bold and center-aligned
    for cell in [sheet["A1"], sheet["B1"], sheet["C1"]]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add some sample data (you can replace this with actual query results)
    sample_data = [
        (101, "09:00", "17:00"),
        (102, "10:00", "18:00"),
    ]

    # Fill the sample data into the sheet
    for row_num, data in enumerate(sample_data, start=2):  # Starting from row 2
        sheet[f"A{row_num}"] = data[0]
        sheet[f"B{row_num}"] = datetime.strptime(data[1], "%H:%M").time()
        sheet[f"C{row_num}"] = datetime.strptime(data[2], "%H:%M").time()

    # Center align the cells in the data section
    for col in "ABC":
        for row in range(2, len(sample_data) + 2):  # Starting from row 2
            cell = sheet[f"{col}{row}"]
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Set the response headers for downloading the file
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="attendance_sample.xlsx"'

    # Save the workbook to the response
    workbook.save(response)

    # Return the response which will trigger the file download
    return response


def get_comp_sites(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()
    
    try:
        user_id = request.session.get('user_id', '')
        selectedCompany = request.POST.get('selectedCompany','')
        cursor.callproc("stp_get_company_wise_site_names", [user_id,selectedCompany])
        for result in cursor.stored_results():
            companywise_site_names = list(result.fetchall())

    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        cursor.callproc("stp_error_log", [fun, str(e), request.user.id])
        print(f"error: {e}")
        return JsonResponse({'result': 'fail', 'message': 'something went wrong!'}, status=500)
    
    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()

    return JsonResponse({'companywise_site_names': companywise_site_names}, status=200)


def save_attendance(request):
    if request.method == 'POST':
        Db.closeConnection()
        m = Db.get_connection()
        cursor = m.cursor()

        try:
            # Get data from POST request
            id = request.POST.get('id')
            employee_id = request.POST.get('employee_id')
            attendance_in = request.POST.get('attendance_in')
            attendance_out = request.POST.get('attendance_out')

            # Call stored procedure to save attendance
            cursor.callproc("stp_save_attendance", [id, employee_id, attendance_in, attendance_out])

        except Exception as e:
            cursor.callproc("stp_error_log", ["save_attendance", str(e), request.user.id])
            return JsonResponse({"result": "fail", "message": "Error saving attendance"}, status=500)

        finally:
            cursor.close()
            m.commit()
            m.close()
            Db.closeConnection()

        return JsonResponse({"result": "success", "message": "Attendance saved successfully"})

    return JsonResponse({"result": "fail", "message": "Invalid request method"}, status=400)

# Function to call the stored procedure and fetch filtered data
def filter_attendance_data(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()
    if request.method == 'GET':
        company_id = request.GET.get('company_id')
        site_id = request.GET.get('site_id')
        shift_date = request.GET.get('shift_date')

        cursor.callproc("stp_filterdateattendance_view", [company_id,site_id,shift_date])
        for result in cursor.stored_results():
            data = list(result.fetchall())
        formatted_data = [
                    {
                      
                        "employee_id": row[0],
                        "employee_name": row[1],
                        "company_id": row[2],
                        "worksite": row[3],
                        "shift_time": row[4],
                        "attendance_in": row[5],
                        "attendance_out": row[6],
                        "id_edit": row[7],
                    }
                    for row in data
        ]

        return JsonResponse({'data': formatted_data})

    return JsonResponse({'error': 'Invalid request method'}, status=400)

@login_required
def check_upload_existence(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        company_id = data.get('company_id')
        site_id = data.get('site_id')
        date_id = data.get('date_id')

        # Check if the combination already exists in the database
        existing_upload = sc_roster.objects.filter(
            company_id=company_id,
            site_id=site_id,
            date=date_id
        ).exists()

        return JsonResponse({'exists': existing_upload})

@login_required
def attendance_upload_excel(request):
    Db.closeConnection()  # Close previous DB connections
    m = Db.get_connection()  # Get a new DB connection
    cursor = m.cursor()
    
    if request.method == 'POST' and request.FILES.get('excelFile'):
        try:
            excel_file = request.FILES.get('excelFile')

            # Check if file is provided
            if not excel_file:
                messages.error(request, {'message': 'No file uploaded', 'status': 'error'})
                return redirect('/attendance_masters?entity=au&type=err')

            # Read the Excel file into a DataFrame
            df = pd.read_excel(excel_file)  # Adjust header to the correct row index
            file_name = excel_file.name
            total_rows = len(df)
            update_count = error_count = 0
            checksum_id = None

            # Initialize user for logging purposes
            global user
            user = request.session.get('user_id', '')

            if df.empty:
                messages.error(request, {'message': 'Uploaded Excel file is empty', 'status': 'error'})
                return redirect('/attendance_masters?entity=au&type=err')

            company_id = request.POST.get('company_id')
            worksite = request.POST.get('site_id')
            shift_date = request.POST.get('date_id')

            # Validate the input fields
            if not (company_id and worksite and shift_date):
                messages.error(request, {'message': 'Missing required fields', 'status': 'error'})
                return redirect('/attendance_masters?entity=au&type=err')

            # Generate checksum_id before processing rows
            cursor.callproc('stp_insert_checksum', ('attendance_upload', company_id, str(datetime.now().month), str(datetime.now().year), file_name))
            for result in cursor.stored_results():
                c = list(result.fetchall())
            checksum_id = c[0][0]  # Get checksum_id after insertion

            # Loop through each row in the DataFrame
            for index, row in df.iterrows():
                employee_id = row.get('Employee ID', None)
                attendance_in = row.get('Attendance In (24 hours format)', None)
                attendance_out = row.get('Attendance Out (24 hours format)', None)

                # Check for missing fields in the current row
                if not (attendance_in and attendance_out and employee_id):
                    error_count += 1
                    continue

                # Check if the combination of company, site, and attendance_date exists in the database
                record_exists = sc_roster.objects.filter(
                    company=company_id,
                    worksite=worksite,
                    shift_date=shift_date
                ).exists()

                if not record_exists:
                    error_count += 1
                    print(f"Error: No record found for company: {company_id}, site: {worksite}, date: {shift_date} at row {index + 1}")
                    continue

                # Check if the employee belongs to the correct company and site
                employee_belongs = sc_roster.objects.filter(
                    employee_id=employee_id,
                    company=company_id,
                    worksite=worksite
                ).exists()

                if not employee_belongs:
                    error_message = f"Employee ID {employee_id} does not belong to company and worksite"
                    cursor.callproc('stp_attendance_error_log', [
                        'attendance_upload',
                        company_id,
                        worksite,
                        file_name,
                        datetime.now().date(),
                        error_message,
                        checksum_id
                    ])
                    error_count += 1
                    continue

                # Add the current datetime dynamically
                attendance_uploaded_date = datetime.now()

                # Prepare parameters for the stored procedure call
                params = [
                    company_id,
                    worksite,
                    shift_date,
                    employee_id,
                    attendance_in,
                    attendance_out,
                    attendance_uploaded_date  # Pass the current datetime
                ]

                # Call the stored procedure
                cursor.callproc('stp_attendance_update', params)

                # Process the result of the stored procedure
                try:
                    for result in cursor.stored_results():
                        r = list(result.fetchall())

                        if r[0][0] != "update":
                            cursor.callproc('stp_attendance_error_log', [
                                'attendance_upload',
                                company_id,
                                '',
                                file_name,
                                datetime.now().date(),
                                str(r[0][0]),
                                checksum_id
                            ])
                            error_count += 1
                        else:
                            update_count += 1

                except Exception as e:
                    print(f"Error occurred in result processing: {str(e)}")

            # Prepare and return the response message
            checksum_msg = f"Total Rows Processed: {total_rows}" f"{f', Updates: {update_count}' if update_count > 0 else ''}" f"{f', Errors: {error_count}' if error_count > 0 else ''}"
            cursor.callproc('stp_update_checksum', (
                'attendance_upload',
                company_id,
                '',
                str(datetime.now().month),
                str(datetime.now().year),
                file_name,
                checksum_msg,
                error_count,
                update_count,
                checksum_id
            ))
            if error_count == 0 and update_count == 0:
                messages.success(request, f"All data uploaded successfully!")
            elif error_count == 0 and update_count > 0:
                messages.success(request, f"All data updated successfully!")
            else:
                messages.warning(request, f"The upload processed {total_rows} rows" f"{f', {update_count} updates' if update_count > 0 else ''}" f", and {error_count} errors; please check the error logs for details.")

        except Exception as e:
            tb = traceback.extract_tb(e.__traceback__)
            fun = tb[0].name
            cursor.callproc("stp_error_log", [fun, str(e), user])
            messages.error(request, 'Oops...! Something went wrong!')

        finally:
            return redirect('/attendance_masters?entity=au&type=err')
        

# palavee Changes

def filter_roster_data(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()
    if request.method == 'GET':
        company_id = request.GET.get('company_id')
        site_id = request.GET.get('site_id')
        shift_date = request.GET.get('shift_date')

        cursor.callproc("stp_filter_roster_data", [company_id,site_id,shift_date])
        for result in cursor.stored_results():
            data = list(result.fetchall())
        formatted_data = [
                    {
                      
                        "employee_id": row[0],
                        "employee_name": row[1],
                        "company_id": row[2],
                        "worksite": row[3],
                        "shift_date": row[4],
                        "uploaded_date": row[5],
                        "id_edit": row[6]
                    }
                    for row in data
        ]

        return JsonResponse({'data': formatted_data})

    return JsonResponse({'error': 'Invalid request method'}, status=400)

@login_required   
def get_worksites(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()
    
    try:
        user_id = request.session.get('user_id', '')
        selectedCompany = request.POST.get('selectedCompany','')
        cursor.callproc("stp_get_slot_siteName", [user_id,selectedCompany])
        for result in cursor.stored_results():
            companywise_site_names = list(result.fetchall())

    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        cursor.callproc("stp_error_log", [fun, str(e), request.user.id])
        print(f"error: {e}")
        return JsonResponse({'result': 'fail', 'message': 'something went wrong!'}, status=500)
    
    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()
    
    return JsonResponse({'companywise_site_names': companywise_site_names}, status=200)

    